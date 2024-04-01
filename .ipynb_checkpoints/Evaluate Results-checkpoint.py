import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import statsmodels.api as sm
from scipy.stats import pearsonr, shapiro
from statsmodels.stats.diagnostic import het_breuschpagan, het_white
from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.stats.stattools import durbin_watson

# Load the workbook
workbook = openpyxl.load_workbook('Study Results.xlsx')

# Get the sheet
processed_data_sheet = workbook['Valid Data']

#number of respondants
total_records = processed_data_sheet.max_row - 1


# Initialize a dictionary to store response counts
gender_dict = {1: "Male", 2: "Female", 3: "Other"}
age_dict = {1: "21&Under", 2: "22-34", 3: "35-44", 4: "45-54", 5: "55-64", 6: "65&Over"}
education_dict = {1: "NoFormal", 2: "GCSE", 3: "A-level", 4: "Bachelors", 5: "Masters", 6: "Doctoral"}
expertise_dict = {1: "Limited", 2: "Basic", 3: "Competent", 4: "Skilled", 5: "Mastery"}

# Function to count the distribution for a specific category
def count_distribution(category_column, category_dict):
    return {id: sum(1 for row in processed_data_sheet.iter_rows(min_row=2, max_col=processed_data_sheet.max_column) if row[category_column - 1].value == id) for id in category_dict}

# Count distribution for each category
gender_counts = count_distribution(2, gender_dict)
age_counts = count_distribution(3, age_dict)
education_counts = count_distribution(4, education_dict)
expertise_counts = count_distribution(5, expertise_dict)


#calculate stats for personality
def calculate_stats(column_index):
    values = [row[column_index - 1].value for row in processed_data_sheet.iter_rows(min_row=2, max_col=processed_data_sheet.max_column)]
    mean_value = np.mean(values)
    std_deviation = np.std(values)
    smallest_value = min(values)
    largest_value = max(values)
    return mean_value, std_deviation, smallest_value, largest_value

extraversion_stats = calculate_stats(6)
agreeableness_stats = calculate_stats(7)
conscientiousness_stats = calculate_stats(8)
negative_emotionality_stats = calculate_stats(9)
open_mindedness_stats = calculate_stats(10)

positive_GAAIS_stats = calculate_stats(11)
negative_GAAIS_stats = calculate_stats(12)

# Close the workbook
workbook.close()



# Print the Demographic results
print("Total Participants:", total_records)

print("\nGender distribution:")
for id, label in gender_dict.items():
    count = gender_counts[id]
    percentage = (count / total_records) * 100
    print(f"{label}: {count} ({percentage:.2f}%)")

print("\nAge distribution:")
for id, label in age_dict.items():
    count = age_counts[id]
    percentage = (count / total_records) * 100
    print(f"{label}: {count} ({percentage:.2f}%)")
    
print("\nEducation distribution:")
for id, label in education_dict.items():
    count = education_counts[id]
    percentage = (count / total_records) * 100
    print(f"{label}: {count} ({percentage:.2f}%)")

print("\nExpertise distribution:")
for id, label in expertise_dict.items():
    count = expertise_counts[id]
    percentage = (count / total_records) * 100
    print(f"{label}: {count} ({percentage:.2f}%)")


# Print Extraversion statistics
print("\nExtraversion statistics:")
print(f"Mean: {extraversion_stats[0]:.2f}")
print(f"Standard Deviation: {extraversion_stats[1]:.2f}")
print(f"Range: {extraversion_stats[2]:.2f} - {extraversion_stats[3]:.2f}")

# Print Agreeableness statistics
print("\nAgreeableness statistics:")
print(f"Mean: {agreeableness_stats[0]:.2f}")
print(f"Standard Deviation: {agreeableness_stats[1]:.2f}")
print(f"Range: {agreeableness_stats[2]:.2f} - {agreeableness_stats[3]:.2f}")

# Print Conscientiousness statistics
print("\nConscientiousness statistics:")
print(f"Mean: {conscientiousness_stats[0]:.2f}")
print(f"Standard Deviation: {conscientiousness_stats[1]:.2f}")
print(f"Range: {conscientiousness_stats[2]:.2f} - {conscientiousness_stats[3]:.2f}")

# Print Negative Emotionality statistics
print("\nNegative Emotionality statistics:")
print(f"Mean: {negative_emotionality_stats[0]:.2f}")
print(f"Standard Deviation: {negative_emotionality_stats[1]:.2f}")
print(f"Range: {negative_emotionality_stats[2]:.2f} - {negative_emotionality_stats[3]:.2f}")

# Print Open-Mindedness statistics
print("\nOpen-Mindedness statistics:")
print(f"Mean: {open_mindedness_stats[0]:.2f}")
print(f"Standard Deviation: {open_mindedness_stats[1]:.2f}")
print(f"Range: {open_mindedness_stats[2]:.2f} - {open_mindedness_stats[3]:.2f}")

# Print Positive GAAIS statistics
print("\nPositive GAAIS statistics:")
print(f"Mean: {positive_GAAIS_stats[0]:.2f}")
print(f"Standard Deviation: {positive_GAAIS_stats[1]:.2f}")
print(f"Range: {positive_GAAIS_stats[2]:.2f} - {positive_GAAIS_stats[3]:.2f}")

# Print Negative GAAIS statistics
print("\nNegative GAAIS statistics:")
print(f"Mean: {negative_GAAIS_stats[0]:.2f}")
print(f"Standard Deviation: {negative_GAAIS_stats[1]:.2f}")
print(f"Range: {negative_GAAIS_stats[2]:.2f} - {negative_GAAIS_stats[3]:.2f}")






#Pearson's
def pearsons_correlation(column_x_index, column_y_index):
    x_values = [row[column_x_index - 1].value for row in processed_data_sheet.iter_rows(min_row=2, max_col=processed_data_sheet.max_column)]
    y_values = [row[column_y_index - 1].value for row in processed_data_sheet.iter_rows(min_row=2, max_col=processed_data_sheet.max_column)]

    correlation_coefficient, p_value = pearsonr(x_values, y_values)
    if p_value < 0.05:
        is_significant = True
    else:
         is_significant = False
        
    if correlation_coefficient < 0:
        correlation_direction = "Negative"
    else:
        correlation_direction = "Positive"


    column_x_title = get_column_title(column_x_index)
    column_y_title = get_column_title(column_y_index)
    
    print(f"\nPearson's correlation coefficient between '{column_x_title}' and '{column_y_title}': {correlation_coefficient}")
    print(f"P-value: {p_value}")
    print(f"Significant: {is_significant}")
       
    if is_significant:
        return (f"'{column_x_title}' and '{column_y_title}'", correlation_direction)
    else:
        return None

def get_column_title(column_index):
    return processed_data_sheet.cell(row=1, column=column_index).value


y_column_indices = [14,19,15,20,16,21,17,22]
x_column_indices = [2,3,4,5,6,7,8,9,10,11,12]
significant_correlations = []

for column_x in x_column_indices:
    for column_y in y_column_indices:
        correlation = pearsons_correlation(column_x, column_y)
        if correlation:
            significant_correlations.append(correlation)

print("\nSignificant Correlations:")
for correlation, direction in significant_correlations:
    print(f"{correlation} ({direction} correlation)")






def multiple_regression(dependent_column_index, independent_column_indices):
    # Extract values from the specified columns
    y_values = [row[dependent_column_index - 1].value for row in processed_data_sheet.iter_rows(min_row=2, max_col=processed_data_sheet.max_column)]
    x_values = [[row[i - 1].value for i in independent_column_indices] for row in processed_data_sheet.iter_rows(min_row=2, max_col=processed_data_sheet.max_column)]

    # Add constant term for intercept
    x_values = sm.add_constant(x_values)

    # Perform multiple regression analysis
    model = sm.OLS(y_values, x_values)
    results = model.fit()

    # Calculate residuals
    residuals = results.resid

    # Shapiro-Wilk test for normality of residuals
    sw_statistic, sw_p_value = shapiro(residuals)  # Perform Shapiro-Wilk test
    print("\nShapiro-Wilk Test for Normality of Residuals:")
    print(f"Shapiro-Wilk statistic: {sw_statistic}")
    print(f"P-value: {sw_p_value}")

    # Breusch-Pagan test for homoskedasticity
    bp_value, bp_p_value, _, _ = het_breuschpagan(residuals, x_values)
    print("\nBreusch-Pagan Test for Homoskedasticity:")
    print(f"Breusch-Pagan statistic: {bp_value}")
    print(f"P-value: {bp_p_value}")

    # Variance Inflation Factor (VIF) for collinearity
    vif_values = [variance_inflation_factor(x_values, i) for i in range(x_values.shape[1])]
    print("\nVariance Inflation Factor (VIF) for Collinearity:")
    for i, vif in enumerate(vif_values):
        if i == 0:
            print("Intercept:", vif)
        else:
            print(f"X{i}:", vif)

    # Durbin-Watson statistic for autocorrelation
    dw_statistic = durbin_watson(residuals)
    print("\nDurbin-Watson Statistic for Autocorrelation:")
    print(f"Durbin-Watson statistic: {dw_statistic}")

    return results

print(f"\n\nChatbot A Quality: {(multiple_regression(17, [2,3,4,5,6,7,8,9,10,11,12])).summary()}")
print(f"\n\nChatbot B Quality: {(multiple_regression(22, [2,3,4,5,6,7,8,9,10,11,12])).summary()}")

print(f"\n\nChatbot A Useful: {(multiple_regression(14, [2,3,4,5,6,7,8,9,10,11,12])).summary()}")
print(f"\n\nChatbot B Useful: {(multiple_regression(19, [2,3,4,5,6,7,8,9,10,11,12])).summary()}")
