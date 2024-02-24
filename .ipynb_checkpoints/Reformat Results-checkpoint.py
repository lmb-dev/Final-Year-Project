import openpyxl
from openpyxl.styles import PatternFill

# Define red and green fill pattern
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

def process_data(values):
    # Extract values from the list and do calculations
    id = values[0]
    gender = values[1]
    age = values[2]
    education = values[3]
    computer = values[4]
    
    extraversion = values[10] + values[15] + (6-values[5])
    agreeableness = values[6] + values[16] + (6-values[11])
    conscientiousness = values[17] + (6-values[7]) + (6-values[12])
    negativeEmotionality = values[8] + values[13] + (6-values[18])
    openMindedness = values[9] + values[19] + (6-values[14])
    
    positiveGAAIS = sum(values[20:24]) / 4
    negativeGAAIS = sum(values[24:28]) / 4 

    # conditional order of which bot is first
    if values[28] == 0:
        bot0msg = values[29]
        bot0use = values[30]
        bot0eng = values[31]
        bot0trust = values[32]
        bot0qual = values[33]      
        bot1msg = values[35]
        bot1use = values[36]
        bot1eng = values[37]
        bot1trust = values[38]
        bot1qual = values[39]
        
    elif values[28] == 1:
        bot0msg = values[35]
        bot0use = values[36]
        bot0eng = values[37]
        bot0trust = values[38]
        bot0qual = values[39]     
        bot1msg = values[29]
        bot1use = values[30]
        bot1eng = values[31]
        bot1trust = values[32]
        bot1qual = values[33]

    #see if row should be rejected
    if values[29] == 0 or values[35] == 0:
        rejected = 1
    else:
        rejected = 0

    # Return the calculated values
    return [
        id, gender, age, education, computer,
        extraversion, agreeableness, conscientiousness,
        negativeEmotionality, openMindedness,
        positiveGAAIS, negativeGAAIS, 
        bot0msg, bot0use, bot0eng, bot0trust, bot0qual,
        bot1msg, bot1use, bot1eng, bot1trust, bot1qual,
        rejected
    ]


# Load the workbook
workbook = openpyxl.load_workbook('Study Results.xlsx')

# Get the sheets
raw_data_sheet = workbook['Raw Data']
processed_data_sheet = workbook['Processed Data']
valid_data_sheet = workbook['Valid Data']

# Clear existing data in the "Processed Data" sheet and "Valid Data" sheet
processed_data_sheet.delete_rows(2, processed_data_sheet.max_row)
valid_data_sheet.delete_rows(2, valid_data_sheet.max_row)

# Iterate through rows starting from the second row (index 2)
for row_index in range(2, raw_data_sheet.max_row + 1):
    # Check if all values in the row are None, indicating an empty row
    if all(
        raw_data_sheet.cell(row=row_index, column=col_index).value is None 
        for col_index in range(1, raw_data_sheet.max_column + 1)
    ):
        # Break out of the loop if an empty row is encountered
        break

    # Get the input data from the "Raw Data" sheet and convert to integers
    input_data = [
        int(raw_data_sheet.cell(row=row_index, column=col_index).value) 
        if raw_data_sheet.cell(row=row_index, column=col_index).value is not None
        else 0  # or replace 0 with the default value you want
        for col_index in range(1, raw_data_sheet.max_column + 1)
    ]

    # Process the data using the process_data function
    result = process_data(input_data)

    # Append the processed data to the "Processed Data" sheet without the fill color
    processed_data_sheet.append(result)

    # Get the cell in the final column
    final_column_cell = processed_data_sheet.cell(row=row_index, column=processed_data_sheet.max_column)

    # Check if the value in the final column is 1 and apply red fill, otherwise apply green fill
    if result[-1] == 1:
        final_column_cell.fill = red_fill
    else:
        final_column_cell.fill = green_fill
        valid_data_sheet.append(result[:-1])            #copy sheet with only accpeted results


# Save the workbook with the processed data
workbook.save('Study Results.xlsx')

# Close the workbook
workbook.close()








